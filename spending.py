import imapclient
import pyzmail
import pprint
import re 
import pandas as pd
import matplotlib.pyplot as plt 
import seaborn as sns
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image

imapObj = imapclient.IMAPClient('imap.gmail.com',ssl=True)

imapObj.login('myemail', 'mypassword')

imapObj.select_folder('INBOX', readonly=False)
ids = imapObj.search(['FROM', 'capitalone@notification.capitalone.com','SUBJECT',"A new transaction was charged to your account"])

# create df and clean/extract data
# ids = ids[0:200]
dates = []
stores = []
amounts = []
for id in ids:
	messages = imapObj.fetch(id,['BODY[]'])
	# pprint.pprint(messages)
	message = pyzmail.PyzMessage.factory(messages[id][b'BODY[]'])
	message_text = message.text_part.get_payload().decode(message.text_part.charset)
	try:
		date = re.search(r"(\d+/\d+/\d+)", message_text).group()
		store = re.search(r"\b(at) (.+),", message_text).group(2)
		amount = re.search(r"\$\d+.\d+", message_text).group()
		
		dates.append(date)
		stores.append(store)
		amounts.append(amount)
	except:
		pass

df_data = {'date': dates,
		   'store': stores,
		   'amount': amounts}

df = pd.DataFrame(df_data)

df['date'] = df['date'].astype('datetime64')
df['amount'] = df['amount'].str.lstrip('$').astype('float64')
df['store'] = df['store'].str.lower()
df['year'] = df['date'].dt.year
df['month'] = df['date'].dt.month
df['week'] = df['date'].dt.week
df['dayofweek'] = df['date'].dt.weekday
df.set_index('date', inplace=True)

wb = Workbook()
wb.save(filename = 'spending.xlsx')

with pd.ExcelWriter('spending.xlsx', engine='openpyxl', mode='a') as writer:
	df.to_excel(writer, header=True, sheet_name='data')

# create visualizations to add to excel spreadsheet
fig, axes = plt.subplots(5, 1, figsize=(12,15))

segments = ['date', 'week','dayofweek','month', 'store']

plt.rcParams['axes.prop_cycle'] = plt.cycler(color=plt.cm.Set3.colors)
colors = plt.rcParams['axes.prop_cycle']()

for i, ax in enumerate(axes.flatten()):
	c = next(colors)['color']
	segment = segments[i]

	if segment == 'date':
		df.groupby(segment).agg('sum').loc[:,'amount'].plot(ax=axes[i], color=c, linewidth=2)
	elif segment == 'store':
		df.groupby(segment).agg('sum').loc[:,'amount'].sort_values(ascending=False).head(20).plot(kind='bar',ax=axes[i], legend=False)
	else:
		df.groupby(segment).agg('mean').loc[:,'amount'].plot(kind='bar', color=c, ax=axes[i])

plt.tight_layout()

plt.savefig('trends.png')

workbook = load_workbook('spending.xlsx')
ws = workbook['data']
img = Image('trends.png')
ws.add_image(img, 'L1')

workbook.save('spending.xlsx')

print('success')
imapObj.delete_messages(ids)
	
